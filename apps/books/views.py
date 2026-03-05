from django.views.generic import ListView, CreateView, UpdateView, DetailView
from django.urls import reverse_lazy, reverse
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Q
from .models import Product, ProductBarcode, ProductUnit, Category
from apps.accounts.decorators import role_required
from apps.accounts.models import ActivityLog
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.http import require_http_methods

from django.utils.decorators import method_decorator
from django.db import transaction, models
import openpyxl
import json
from io import BytesIO

class ProductListView(LoginRequiredMixin, ListView):
    model = Product
    template_name = 'books/product_list.html'
    context_object_name = 'products'
    paginate_by = 20

    def get_queryset(self):
        query = self.request.GET.get('q')
        if query:
            return Product.objects.filter(name__icontains=query)
        return Product.objects.all()

@method_decorator(role_required(allowed_roles=['admin']), name='dispatch')
class ProductCreateView(LoginRequiredMixin, CreateView):
    model = Product
    template_name = 'books/product_form.html'
    fields = ['product_type', 'name', 'author', 'page_count', 'language', 'description', 'categories', 'base_unit',
              'has_package', 'package_type', 'package_qty',
              'purchase_price', 'selling_price', 'tax_rate', 'discount_rate',
              'package_purchase_price', 'package_selling_price', 'package_tax_rate', 'package_discount_rate',
              'min_price', 'min_stock_level', 'image', 'is_active']
    success_url = reverse_lazy('product_list')

    def form_valid(self, form):
        with transaction.atomic():
            response = super().form_valid(form)
            self.object.created_by = self.request.user
            self.object.save()
            
            # Handle Barcodes
            new_barcodes = self.request.POST.getlist('new_barcodes[]')
            for bc in new_barcodes:
                if bc.strip():
                    ProductBarcode.objects.get_or_create(
                        product=self.object,
                        barcode=bc.strip(),
                        defaults={'is_primary': False}
                    )

            # Handle Units
            unit_names = self.request.POST.getlist('unit_name[]')
            unit_factors = self.request.POST.getlist('unit_factor[]')
            unit_prices = self.request.POST.getlist('unit_price[]')
            unit_barcodes = self.request.POST.getlist('unit_barcode[]')
            
            for i in range(len(unit_names)):
                if unit_names[i].strip():
                    ProductUnit.objects.get_or_create(
                        product=self.object,
                        name=unit_names[i].strip(),
                        defaults={
                            'conversion_factor': unit_factors[i] or 1,
                            'selling_price': unit_prices[i] or self.object.selling_price,
                            'barcode': unit_barcodes[i] if i < len(unit_barcodes) else ''
                        }
                    )

            ActivityLog.objects.create(
                user=self.request.user,
                action="create",
                action_description=f"تم إضافة منتج جديد: {self.object.name}",
                object_type="Product",
                object_id=self.object.product_id
            )
            return response

@role_required(allowed_roles=['admin'])
def export_products_excel(request):
    products = Product.objects.all()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Products'
    
    headers = ['ID', 'Name', 'Barcode', 'Price', 'Stock']
    sheet.append(headers)
    
    for p in products:
        primary_barcode = p.barcodes.filter(is_primary=True).first() or p.barcodes.first()
        barcode_str = primary_barcode.barcode if primary_barcode else ''
        sheet.append([p.product_id, p.name, barcode_str, p.selling_price, p.current_stock])
    
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=products.xlsx'
    return response

@role_required(allowed_roles=['admin'])
def import_products_excel(request):
    if request.method == 'POST' and request.FILES.get('file'):
        excel_file = request.FILES['file']
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not any(row): continue
            
            # التوافق مع الصيغة الجديدة (ID, Name, Barcode, Price, Stock) 
            # أو القديمة (Name, Barcode, Price, Stock)
            if len(row) >= 5:
                p_id, name, barcode, price, stock = row[:5]
            else:
                p_id = None
                name, barcode, price, stock = row[:4]

            if p_id:
                product, created = Product.objects.update_or_create(
                    product_id=p_id,
                    defaults={'name': name, 'selling_price': price, 'current_stock': stock}
                )
            else:
                product, created = Product.objects.get_or_create(
                    name=name,
                    defaults={'selling_price': price, 'current_stock': stock}
                )
            if barcode:
                ProductBarcode.objects.get_or_create(product=product, barcode=barcode, defaults={'is_primary': True})
        messages.success(request, "تم استيراد/تحديث المنتجات بنجاح")
        ActivityLog.objects.create(
            user=request.user,
            action="استيراد منتجات",
            details=f"تم استيراد/تحديث المنتجات من ملف Excel: {excel_file.name}"
        )
        return redirect('product_list')
    return render(request, 'books/import.html')

@method_decorator(role_required(allowed_roles=['admin']), name='dispatch')
class ProductUpdateView(LoginRequiredMixin, UpdateView):
    model = Product
    template_name = 'books/product_form.html'
    fields = ['product_type', 'name', 'author', 'page_count', 'language', 'description', 'categories', 'base_unit',
              'has_package', 'package_type', 'package_qty',
              'purchase_price', 'selling_price', 'tax_rate', 'discount_rate',
              'package_purchase_price', 'package_selling_price', 'package_tax_rate', 'package_discount_rate',
              'min_price', 'min_stock_level', 'image', 'is_active']
    success_url = reverse_lazy('product_list')

    def get_object(self, queryset=None):
        # product_id في الـ URL
        return get_object_or_404(Product, product_id=self.kwargs.get('product_id'))

    def form_valid(self, form):
        with transaction.atomic():
            response = super().form_valid(form)
            self.object.updated_by = self.request.user
            self.object.save()

            # Handle New Barcodes
            new_barcodes = self.request.POST.getlist('new_barcodes[]')
            for bc in new_barcodes:
                if bc.strip():
                    ProductBarcode.objects.get_or_create(
                        product=self.object,
                        barcode=bc.strip(),
                        defaults={'is_primary': False}
                    )

            # Handle New Units
            unit_names = self.request.POST.getlist('unit_name[]')
            unit_factors = self.request.POST.getlist('unit_factor[]')
            unit_prices = self.request.POST.getlist('unit_price[]')
            unit_barcodes = self.request.POST.getlist('unit_barcode[]')
            
            for i in range(len(unit_names)):
                if unit_names[i].strip():
                    ProductUnit.objects.update_or_create(
                        product=self.object,
                        name=unit_names[i].strip(),
                        defaults={
                            'conversion_factor': unit_factors[i] or 1,
                            'selling_price': unit_prices[i] or self.object.selling_price,
                            'barcode': unit_barcodes[i] if i < len(unit_barcodes) else ''
                        }
                    )

            ActivityLog.objects.create(
                user=self.request.user,
                action="update",
                action_description=f"تم تحديث بيانات المنتج: {self.object.name}",
                object_type="Product",
                object_id=self.object.product_id
            )
            return response

@login_required
def product_search_ajax(request):
    query = request.GET.get('q', '')
    if len(query) < 2:
        return JsonResponse({'results': []})
    
    # Search by name or any of its barcodes
    products = Product.objects.filter(
        Q(name__icontains=query) | 
        Q(product_id__icontains=query) |
        Q(barcodes__barcode__icontains=query)
    ).distinct()[:10]
    
    results = []
    for p in products:
        primary_barcode = p.barcodes.filter(is_primary=True).first() or p.barcodes.first()
        results.append({
            'id': p.id,
            'product_id': p.product_id,
            'name': p.name,
            'price': float(p.selling_price),
            'stock': float(p.current_stock),
            'barcode': primary_barcode.barcode if primary_barcode else '',
            'unit': p.base_unit
        })
    
    return JsonResponse({'results': results})

@login_required
def api_get_product_by_barcode(request):
    barcode = request.GET.get('barcode', '')
    if not barcode:
        return JsonResponse({'success': False, 'error': 'No barcode provided'})
    
    # Check regular barcodes and unit barcodes
    p_barcode = ProductBarcode.objects.filter(barcode=barcode).select_related('product').first()
    if p_barcode:
        product = p_barcode.product
    else:
        # Check units
        unit = ProductUnit.objects.filter(barcode=barcode).select_related('product').first()
        if unit:
            product = unit.product
            # Return unit info as well? For POS, we might want the unit price
        else:
            return JsonResponse({'success': False, 'error': 'Product not found'})
    
    return JsonResponse({
        'success': True,
        'id': product.id,
        'product_id': product.product_id,
        'name': product.name,
        'price': float(product.selling_price),
        'stock': float(product.current_stock),
        'unit': product.base_unit
    })

@role_required(allowed_roles=['admin'])
def product_delete(request, product_id):
    product = get_object_or_404(Product, product_id=product_id)
    if product.has_invoices():
        messages.error(request, f"لا يمكن حذف المنتج {product.name} لوجود فواتير مرتبطة به.")
        return redirect('product_list')
    
    product_name = product.name
    product_id_val = product.product_id
    product.delete()
    
    ActivityLog.objects.create(
        user=request.user,
        action="delete",
        action_description=f"تم حذف المنتج: {product_name}",
        object_type="Product",
        object_id=product_id_val
    )
    messages.success(request, f"تم حذف المنتج {product_name} بنجاح.")
    return redirect('product_list')

@login_required
def api_search_categories(request):
    query = request.GET.get('q', '').strip()
    exclude_ids = request.GET.get('exclude', '').split(',')
    exclude_ids = [int(eid) for eid in exclude_ids if eid.isdigit()]

    if query:
        categories = Category.objects.filter(name__icontains=query)
    else:
        categories = Category.objects.all()
        
    if exclude_ids:
        categories = categories.exclude(id__in=exclude_ids)
        
    categories = categories[:10]
        
    results = [{'id': c.id, 'name': c.name} for c in categories]
    return JsonResponse({'results': results})

@login_required
@require_http_methods(["POST"])
def api_create_category(request):
    try:
        data = json.loads(request.body)
        name = data.get('name', '').strip()
        
        if not name:
            return JsonResponse({'success': False, 'error': 'اسم  مطلوب'})
            
        category, created = Category.objects.get_or_create(
            name=name
        )
        return JsonResponse({
            'success': True, 
            'category': {
                'id': category.id, 
                'name': category.name
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def api_author_suggestions(request):
    """API Autocomplete للمؤلفين المستخدمين سابقاً"""
    query = request.GET.get('q', '').strip()
    qs = Product.objects.exclude(author__isnull=True).exclude(author='')
    if query:
        qs = qs.filter(author__icontains=query)
    authors = list(qs.values_list('author', flat=True).distinct()[:10])
    return JsonResponse({'results': authors})


@login_required
def api_package_types(request):
    """API Autocomplete لأنواع العبوات المستخدمة سابقاً"""
    query = request.GET.get('q', '').strip()
    qs = Product.objects.filter(has_package=True).exclude(package_type='')
    if query:
        qs = qs.filter(package_type__icontains=query)
    types = list(qs.values_list('package_type', flat=True).distinct()[:10])
    return JsonResponse({'results': types})
